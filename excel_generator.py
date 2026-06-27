import pandas as pd
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT_FOLDER = Path("output")


def _ensure_output_folder():
    OUTPUT_FOLDER.mkdir(exist_ok=True)


def _format_excel(file_path):
    from openpyxl import load_workbook

    workbook = load_workbook(file_path)

    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="D9EAF7",
                end_color="D9EAF7",
                fill_type="solid"
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                value = str(cell.value) if cell.value else ""
                max_length = max(max_length, len(value))

            sheet.column_dimensions[column_letter].width = min(
                max_length + 5,
                60
            )

    workbook.save(file_path)


def export_single_sheet(data, file_name, sheet_name):
    _ensure_output_folder()

    file_path = OUTPUT_FOLDER / file_name

    df = pd.DataFrame(data)

    df.to_excel(
        file_path,
        sheet_name=sheet_name,
        index=False
    )

    _format_excel(file_path)

    print(f"\nExcel created: {file_path}")


def export_complete_workbook(workbook_data):
    _ensure_output_folder()

    file_path = OUTPUT_FOLDER / "QA_Workbook.xlsx"

    with pd.ExcelWriter(
        file_path,
        engine="openpyxl"
    ) as writer:

        for sheet_name, data in workbook_data.items():
            df = pd.DataFrame(data)

            df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False
            )

    _format_excel(file_path)

    print(f"\nComplete QA Workbook created: {file_path}")